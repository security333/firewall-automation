import openpyxl
from openpyxl.styles import Font

def load_ip_addresses(filename):
    ip_addresses = set()
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            ip_addresses.add(cell)
    return ip_addresses

def highlight_matches(ip_addresses, input_filename, output_filename):
    workbook = openpyxl.load_workbook(input_filename)
    sheet = workbook.active

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in ip_addresses:
                cell.font = Font(bold=True)

    workbook.save(output_filename)

if __name__ == "__main__":
    ip_addresses_filename = "ip_addresses.xlsx"  # File containing IP addresses
    input_filename = "input.xlsx"  # File to read and highlight matches
    output_filename = "output.xlsx"  # Output file with matches highlighted

    ip_addresses = load_ip_addresses(ip_addresses_filename)
    highlight_matches(ip_addresses, input_filename, output_filename)
